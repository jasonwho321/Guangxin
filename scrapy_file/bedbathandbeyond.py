from concurrent.futures import ProcessPoolExecutor
import pandas as pd
import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import json
# 用于缓存已爬取URL的字典
url_cache = {}
def get_redirected_url(url):
    if url in url_cache:
        return url_cache[url]

    proxyip = "http://storm-maplin_area-US_city-LosAngeles_life-1:Homycasa2012@proxy.stormip.cn:1000"

    proxies = {
        'http': proxyip,
        'https': proxyip,
    }
    ua = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36 Edg/106.0.1370.42'
    headers = {'User-Agent': ua}
    headers['Range'] = 'bytes=0-10'
    retries = 3
    for _ in range(retries):
        try:
            response = requests.get(url, allow_redirects=True, proxies=proxies, headers=headers)
            if 'gateway.zscalerthree' not in response.url:
                print(f"Redirected to {response.url}")
                url_cache[url] = response.url  # 缓存这个结果
                return response.url
        except requests.exceptions.RequestException as e:
            print(f"Error occurred: {e}")

    url_cache[url] = None  # 缓存这个结果，即使它是None
    return None

def scrapy_review_with_retry(url):
    retries = 3
    for _ in range(retries):
        try:
            return request_url(url)
        except requests.exceptions.RequestException as e:
            print(f"Error occurred: {e}, Retrying...")
    return None
def clean_data(file_dir = '/Users/huzhang/Library/CloudStorage/坚果云-john.hu@39f.net/「晓望集群」/S数据分析/OS爬虫/1008 TO John-批零比监控逻辑表-GX&Homylin.xlsx',output = '/Users/huzhang/Library/CloudStorage/坚果云-john.hu@39f.net/「晓望集群」/S数据分析/OS爬虫/cleaned_data.csv'):
    # Load the workbook and select the first worksheet
    wb = load_workbook(file_dir)
    ws1 = wb['Homylin账号批零比监控']

    # Iterate through each row to get the hyperlink from a specific column
    # Assume the hyperlink is in column F (6th column), and data starts from row 4
    hyperlinks1 = []
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=6, max_col=6):
        for cell in row:
            print(cell.hyperlink.target if cell.hyperlink else None)
            hyperlinks1.append(cell.hyperlink.target if cell.hyperlink else None)

    # Now, you can replace the 'Product page' column in your DataFrame with these hyperlinks
    df1 = pd.read_excel(
        file_dir,
        sheet_name='Homylin账号批零比监控')

    df1['Product page'] = hyperlinks1

    # 保存新的 DataFrame
    df1.to_csv(output, index=False, encoding='gb18030')
def extract_product_info(html_content):
    # Initialize BeautifulSoup object
    soup = BeautifulSoup(html_content, 'html.parser')

    # Find the script tag containing JSON data
    script_tag = soup.find('script', {'id': '__NEXT_DATA__', 'type': 'application/json'})
    if not script_tag:
        return None

    # Load JSON content into a Python object
    json_content = script_tag.string
    json_data = json.loads(json_content)
    data_dict = {'optionid': [], 'subsku': [], 'decription': [], 'price': [], 'qtyonhand': [], 'full_sku': []}
    # Define a function to recursively search for the key in a nested dictionary or list
    # Search for the target keys in the JSON object to find their paths
    option_list = json_data.get('props', {}).get('pageProps', {}).get('product', {}).get('options')
    meta_full_sku = json_data.get('props', {}).get('pageProps', {}).get('meta', {}).get('dataLayer', {}).get(
        'full_sku')
    if option_list and meta_full_sku:
        for i in option_list:
            optionid = i.get('optionId', '-')
            subsku = i.get('subSku', '-')
            decription = i.get('decription', '-')
            price = i.get('price', '-')
            qtyonhand = i.get('qtyOnHand', '-')
            full_sku = meta_full_sku[:8]+'-'+subsku
            data_dict['optionid'].append(optionid)
            data_dict['subsku'].append(subsku)
            data_dict['decription'].append(decription)
            data_dict['price'].append(price)
            data_dict['qtyonhand'].append(qtyonhand)
            data_dict['full_sku'].append(full_sku)
        print(data_dict)
    else:
        data_dict = {'optionid': [], 'subsku': [], 'decription': [], 'price': [], 'qtyonhand': [],'full_sku':[]}
        pass
    return data_dict
def main():
    df1, df2 = read_data()
    df2 = df2.dropna(subset=['Product page'])
    result_df = pd.DataFrame()

    with ProcessPoolExecutor() as executor:
        list_of_data_dicts = list(executor.map(scrapy_review_with_retry, [row['Product page'] for _, row in df2.iterrows()]))

    for data_dict in list_of_data_dicts:
        data_df = pd.DataFrame(data_dict)
        result_df = pd.concat([result_df, data_df], axis=0)

    result = pd.merge(df1, result_df, left_on='Full SKU', right_on='full_sku')
    # 计算"零批比"并添加到新的列
    result['零批比'] = result['price'] / result['First Cost']

    # 添加"大于2"列，根据"零批比"的值设置列的内容
    result['大于2'] = result['零批比'].apply(lambda x: '大于2' if x > 2 else '小于2')

    # 筛选出"大于2"的Goods Name，并组成一个列表
    goods_name_list = result[result['大于2'] == '大于2']['Goods Name'].tolist()


    print(f"大于2的商品数量：{str(len(goods_name_list))}")
    # 根据“平台”列的值将DataFrame分为两个不同的DataFrames
    df_homylininc = result.loc[result['平台'] == 'Homylininc']
    df_guangdongguangxin = result.loc[result['平台'] == 'Guangdongguangxin']

    # 将这两个新的DataFrames保存为CSV文件
    df_homylininc.to_csv(
        '/Users/huzhang/Library/CloudStorage/坚果云-john.hu@39f.net/「晓望集群」/S数据分析/OS爬虫/homylininc.csv',
        encoding='gb18030', index=False)
    df_guangdongguangxin.to_csv(
        '/Users/huzhang/Library/CloudStorage/坚果云-john.hu@39f.net/「晓望集群」/S数据分析/OS爬虫/guangdongguangxin.csv',
        encoding='gb18030', index=False)
def read_data():
    # 读取选定列（'Full SKU', 'Goods Name', 'First Cost'）从CSV文件到 DataFrame1
    df1 = pd.read_csv('/Users/huzhang/Library/CloudStorage/坚果云-john.hu@39f.net/「晓望集群」/S数据分析/OS爬虫/cleaned_data.csv', encoding='gb18030')
    df1 = df1.drop_duplicates(subset=['Full SKU'], keep='first').reset_index(drop=True)

    # 读取 'Product page' 列，删除重复项，并将其保存到 DataFrame2
    df2 = pd.read_csv('/Users/huzhang/Library/CloudStorage/坚果云-john.hu@39f.net/「晓望集群」/S数据分析/OS爬虫/cleaned_data.csv', encoding='gb18030',
                      usecols=['Product page']).drop_duplicates().reset_index(drop=True)
    return df1,df2
def request_url(url):
    proxyip = "http://storm-maplin_area-US_city-LosAngeles_life-1:Homycasa2012@proxy.stormip.cn:1000"

    proxies = {
        'http': proxyip,
        'https': proxyip,
    }
    ua = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36 Edg/106.0.1370.42'
    headers = {'User-Agent': ua}
    response = requests.get(url,headers=headers,proxies=proxies)
    return extract_product_info(response.text)

if __name__ == '__main__':
    # clean_data('/Users/huzhang/Library/CloudStorage/坚果云-john.hu@39f.net/「晓望集群」/S数据分析/OS爬虫/1012 TO John-批零比监控逻辑表-GX&Homylin.xlsx')
    # main()
    output = r'/Users/huzhang/Library/CloudStorage/坚果云-john.hu@39f.net/「晓望集群」/S数据分析/OS爬虫/cleaned_data.csv'
    df1 = pd.read_csv(output, encoding='gb18030')

    # 使用get_redirected_url函数获取重定向后的URL，并存储在新列中
    df1['Redirected Product page'] = df1['Product page'].apply(get_redirected_url)

    # 重命名原始 'Product page' 列为 'Original page'
    df1.rename(columns={'Product page': 'Original page'}, inplace=True)
    df1.to_csv('/Users/huzhang/Library/CloudStorage/坚果云-john.hu@39f.net/「晓望集群」/S数据分析/OS爬虫/cleaned_data_1.csv', index=False, encoding='gb18030')
    pass