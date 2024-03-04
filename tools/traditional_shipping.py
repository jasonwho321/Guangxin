import os
import openpyxl
import requests
import pandas as pd
from guangxin_base import bot_push_text
from datetime import datetime
file_dir = '/Users/huzhang/Library/CloudStorage/坚果云-john.hu@39f.net/「晓望集群」/S数据分析/传统业务/【晓望梅观】各机构传统客户2023出运金额.xlsx'
def process_response(response, xls_temp_file, csv_output_file,engine="xlrd"):
    try:
        response.raise_for_status()
        with open(xls_temp_file, "wb") as f:
            f.write(response.content)
        if engine == "xlrd":
            df = pd.read_excel(xls_temp_file, engine='xlrd', header=1)
        else:
            df = pd.read_excel(xls_temp_file, engine='openpyxl')
        df.to_csv(csv_output_file, index=False)
        os.remove(xls_temp_file)
        return df
    except requests.HTTPError as http_err:
        bot_push_text(f"HTTP error occurred: {http_err}", ['18062351119'])
    except Exception as e:
        bot_push_text(f"An error occurred: {e}", ['18062351119'])

def download_procurementPlan_transportInfo():
    url = "https://chinawms.houseinbox.com/rbcs/procurementPlan/exportData"
    url_1 = 'https://chinawms.houseinbox.com/rbcs/transport/exportData'
    headers = {'userid': '825'}
    payload = {
        "query": {
            "ainstitutionids": [384,961,962,454,843,332,589,976,977,274,531,536,543,736,929,611,614,998,871,1000,1001,298,490,1004,493,1005,366,494,623,1009,566,633,763,317,765,446]
        }
    }
    payload_1 = {
        "transportQuery": {
            "ainstitutionAids": [384,961,962,454,843,332,589,976,977,274,531,536,543,736,929,611,614,998,871,1000,1001,298,490,1004,493,1005,366,494,623,1009,566,633,763,317,765,446]
        }
    }
    # 发送请求
    response = requests.post(url, json=payload, headers=headers)
    response_1 = requests.post(url_1, json=payload_1, headers=headers)
    # 检查响应状态
    xls_temp_file = "temp.xls"
    csv_output_file = 'China_WMS.csv'
    df1 = process_response(response, xls_temp_file, csv_output_file)

    xls_temp_file_1 = "temp_1.xls"
    csv_output_file_1 = 'China_WMS_2.csv'
    df2 = process_response(response_1, xls_temp_file_1, csv_output_file_1,engine='openpyxl')
    return df1,df2
def process_receivable(df):
    # Convert the '采购计划ETD' column to datetime format
    df['采购计划ETD'] = pd.to_datetime(df['采购计划ETD'], errors='coerce')

    # Filter data for the year 2023
    df_2023 = df[df['采购计划ETD'].dt.year == 2023]

    # Filter the data based on the '单据状态' column with values '新增', '已提交', '合同已签订'
    statuses = ['新增', '提交', '合同已签订']
    df_2023_filtered = df_2023[df_2023['单据状态'].isin(statuses)]

    # Create the '应收金额' column based on '销售总金额(USD)' and '佣金总额(USD)'
    df_2023_filtered.loc[:, '销售总金额(USD)'] = pd.to_numeric(df_2023_filtered['销售总金额(USD)'], errors='coerce')
    df_2023_filtered.loc[:, '扣佣'] = pd.to_numeric(df_2023_filtered['扣佣'], errors='coerce')
    df_2023_filtered.loc[:, '应收金额'] = df_2023_filtered['销售总金额(USD)'] - df_2023_filtered['扣佣']
    # Create a new column with the first three characters of the organization name
    df_2023_filtered.loc[:, '组织_prefix'] = df_2023_filtered['组织'].str[:3]

    # Adjust the '组织_prefix' column to replace '正一观' with '新天地'
    new_tian_di_prefixes = ['正一观', '新天地']
    df_2023_filtered.loc[:, '组织_prefix'] = df_2023_filtered['组织_prefix'].apply(
        lambda x: '新天地' if x in new_tian_di_prefixes else x)

    # 对整个 DataFrame 按照 '组织_prefix' 分组，并对 '应收金额' 列进行求和
    grouped_total = df_2023_filtered.groupby('组织_prefix')['应收金额'].sum().reset_index()

    # 调整'风山涧'的数据
    wind_mountain_initial = 291774.85
    grouped_total.loc[grouped_total['组织_prefix'] == '风山涧', '应收金额'] += wind_mountain_initial

    # 重命名列名
    today = datetime.now()
    current_week = today.isocalendar()[1]
    grouped_total.columns = ["组织_prefix", f"Week {current_week} 在手订单金额（含已出运）"]
    return grouped_total
def process_shipping(df):
    # Convert the '采购计划ETD' column to datetime format
    df['ETD'] = pd.to_datetime(df['ETD'], errors='coerce')

    # Filter data for the year 2023
    df_2023_filtered = df[df['ETD'].dt.year == 2023]

    # Create the '应收金额' column based on '销售总金额(USD)' and '佣金总额(USD)'
    df_2023_filtered.loc[:, '出口收汇USD'] = pd.to_numeric(df_2023_filtered['出口收汇USD'], errors='coerce')
    # Create a new column with the first three characters of the organization name
    df_2023_filtered.loc[:, '组织_prefix'] = df_2023_filtered['业务组'].str[:3]

    # Adjust the '组织_prefix' column to replace '正一观' with '新天地'
    new_tian_di_prefixes = ['正一观', '新天地']
    df_2023_filtered.loc[:, '组织_prefix'] = df_2023_filtered['组织_prefix'].apply(
        lambda x: '新天地' if x in new_tian_di_prefixes else x)

    # Group the data by the updated '组织_prefix' and sum the '应收金额' up to current and previous weeks
    # 对整个 DataFrame 按照 '组织_prefix' 分组，并对 '应收金额' 列进行求和
    grouped_total = df_2023_filtered.groupby('组织_prefix')['出口收汇USD'].sum().reset_index()

    # 调整'风山涧'的数据
    wind_mountain_initial = 291774.85
    grouped_total.loc[grouped_total['组织_prefix'] == '风山涧', '出口收汇USD'] += wind_mountain_initial

    # 重命名列名
    today = datetime.now()
    current_week = today.isocalendar()[1]
    grouped_total.columns = ["组织_prefix", f"Week {current_week} 出口收汇（2023.01.01至今）"]
    return grouped_total
def main():
    df1, df2 = download_procurementPlan_transportInfo()
    grouped_total_receivable = process_receivable(df1)
    grouped_total_shipping = process_shipping(df2)
    # 将df转换为字典
    # 重命名列名
    today = datetime.now()
    current_week = today.isocalendar()[1]
    preview_week = current_week - 1
    order_data = grouped_total_receivable.set_index('组织_prefix')[f'Week {current_week} 在手订单金额（含已出运）'].to_dict()
    export_data = grouped_total_shipping.set_index('组织_prefix')[f'Week {current_week} 出口收汇（2023.01.01至今）'].to_dict()

    # 加载原始Excel文件
    wb = openpyxl.load_workbook(file_dir)
    ws = wb.active

    # 更改表头
    headers = {
        'D1': f'Week {preview_week} 在手订单金额（含已出运）',
        'E1': f'Week {current_week} 在手订单金额（含已出运）',
        'G1': f'Week {preview_week} 出口收汇（2023.01.01至今）',
        'H1': f'Week {current_week} 出口收汇（2023.01.01至今）'
    }
    for cell, header in headers.items():
        ws[cell].value = header

    # 更新 Week 42 数据和 Week 43 数据
    for row in range(2, 9):  # 假设数据从第2行开始，到第8行结束
        group = ws[f'B{row}'].value
        if group in order_data:
            ws[f'D{row}'].value = ws[f'E{row}'].value  # 在手订单金额 Week 42
            ws[f'E{row}'].value = order_data[group]  # 在手订单金额 Week 43
        if group in export_data:
            ws[f'G{row}'].value = ws[f'H{row}'].value  # 出口收汇 Week 42
            ws[f'H{row}'].value = export_data[group]  # 出口收汇 Week 43

    # 获取2-8行的BCDEGH列数据
    data_rows = []
    for row in range(2, 9):
        data_row = [ws.cell(row=row, column=col).value for col in [2, 3, 4, 5, 7, 8]]
        data_rows.append(data_row)

    # 根据“Week 43 在手订单金额（含已出运）”（即D列，列表中的索引为2）排序
    sorted_data_rows = sorted(data_rows, key=lambda x: x[2], reverse=True)

    # 将排序后的数据写回到Excel的相应位置
    for row, data_row in enumerate(sorted_data_rows, start=2):
        for col_index, col in enumerate([2, 3, 4, 5, 7, 8]):
            ws.cell(row=row, column=col, value=data_row[col_index])

    # 保存Excel
    wb.save('/Users/huzhang/Library/CloudStorage/坚果云-john.hu@39f.net/「晓望集群」/S数据分析/传统业务/【晓望梅观】各机构传统客户2023出运金额2.xlsx')

if  __name__ == '__main__':
    main()
    pass
